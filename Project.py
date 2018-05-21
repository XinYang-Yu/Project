from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

results = load_workbook('Result Sheet for N(A) level candidates.xlsx')
result_table = results.active
arranged = results.create_sheet('Arranged')
emb3 = results.create_sheet('EMB3')
emb3.freeze_panes = 'A2'

lookup_table = load_workbook('Lookup_Table.xlsx')
subject_list = lookup_table.active


def arrange():

	# Particulars
	starting_row = int(input("Starting row: "))
	ending_row = int(input("Ending row: "))
	columns = int(input("Number of columns: "))
	output_row = 1

	for row in result_table.iter_rows(min_row = starting_row, max_col = 1, max_row = ending_row):
		for cell in row:
			if type(cell.value) == str:
				arranged.cell(row = output_row, column = 1).value = result_table.cell(row = cell.row, column = 2).value
				arranged.cell(row = output_row, column = 2).value = result_table.cell(row = cell.row, column = 3).value
				output_row += 1

	# Grades
	output_row = 0
	output_column = 4
	current_number = 0

	for i in range(starting_row, ending_row + 1):
		for j in range(5, columns):
			if result_table.cell(row = i, column = j).value:
				# If same number or empty cell
				if result_table.cell(row = i, column = 1).value == current_number or not result_table.cell(row = i, column = 1).value:
					arranged.cell(row = output_row, column = output_column).value = result_table.cell(row = i, column = j).value
				else:  # If different number
					output_column = 4
					output_row += 1
					current_number = result_table.cell(row = i, column = 1).value
					arranged.cell(row = output_row, column = output_column).value = result_table.cell(row = i, column = j).value
				output_column += 1


def subjects():

	# Copying over from Arranged to EMB3
	rows = arranged.max_row
	columns = arranged.max_column

	for row in arranged.iter_rows(min_row = 1, max_col = columns, max_row = rows):
		for cell in row:
			emb3.cell(row = cell.row + 2, column = column_index_from_string(cell.column)).value = cell.value

	# Subjects
	current_column = 12
	done = []

	for i in range(3, rows + 2):
		for j in range(4, columns + 1):
			if emb3.cell(row = i, column = j).value and emb3.cell(row = i, column = j).value[:5] not in done:
				# Subject code
				emb3.cell(row = 1, column = current_column).value = emb3.cell(row = i, column = j).value[:5]
				# Subject category
				for col in subject_list.iter_cols(min_row = 2, max_col = 1, max_row = subject_list.max_row):
					for cell in col:
						if str(cell.value) == emb3.cell(row = i, column = j).value[:5]:
							emb3.cell(row = 2, column = current_column).value = subject_list.cell(row = cell.row, column = 4).value
				current_column += 1
				done.append(emb3.cell(row = i, column = j).value[:5])

	# Grades
	for i in range(3, rows + 2):
		for j in range(4, columns + 1):
			if emb3.cell(row = i, column = j).value:
				for k in range(12, emb3.max_column + 1):
					if emb3.cell(row = 1, column = k).value == emb3.cell(row = i, column = j).value[:5]:
						emb3.cell(row = i, column = k).value = emb3.cell(row = i, column = j).value[6]


def calculation():

	rows = emb3.max_row
	starting_column = emb3.max_column + 2

	#English
	for i in range(arranged.max_column + 2, emb3.max_column + 1):
		pass

arrange()
subjects()
calculation()

results.save('Output.xlsx')
